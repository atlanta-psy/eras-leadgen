#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
диспетчер отправки (безопасный, человек в цикле).
не шлёт сам — формирует дневную пачку с учётом лимитов, прогрева и фоллоу-апов,
ведёт учёт «кому писали / кто ответил».

команды:
  python3 send_manager.py today --niche guest_house     # сформировать пачку на сегодня
  python3 send_manager.py replied --niche guest_house --name "Оливия"   # отметить, что ответили
  python3 send_manager.py skip --niche guest_house --name "Горки"       # исключить лид
  python3 send_manager.py stats --niche guest_house     # воронка
порядок работы: запустил today -> отправил пачку руками -> кто ответил, отметил replied.
"""
from __future__ import annotations
import argparse, csv, json, sys, time, smtplib, ssl
from email.mime.text import MIMEText
from email.utils import formataddr
from datetime import date, datetime
from pathlib import Path
import yaml

ROOT = Path(__file__).resolve().parent
DATA = ROOT / "data"
CONFIG = ROOT / "configs"
CHANNELS = ["email", "telegram", "vk"]   # phone — только звонок, в рассылку не идёт


def today() -> date:
    return date.today()


def dparse(s: str):
    return datetime.strptime(s, "%Y-%m-%d").date() if s else None


def lead_key(r: dict) -> str:
    return (r.get("email") or r.get("telegram") or r.get("vk") or r.get("phone") or r.get("name") or "").strip().lower()


def load_outreach(niche: str) -> dict:
    path = DATA / f"outreach_{niche}.csv"
    rows = {}
    for r in csv.DictReader(path.open(encoding="utf-8-sig")):
        rows[lead_key(r)] = r
    return rows


def state_path(niche): return DATA / f"state_{niche}.csv"
def meta_path(niche): return DATA / f"campaign_{niche}.json"
STATE_COLS = ["key", "name", "channel", "email", "phone", "telegram", "vk", "status", "last_sent", "replied"]


def load_state(niche: str, outreach: dict) -> dict:
    p = state_path(niche)
    if p.exists():
        return {r["key"]: r for r in csv.DictReader(p.open(encoding="utf-8-sig"))}
    # инициализация из очереди
    st = {}
    for k, r in outreach.items():
        if r.get("channel") not in CHANNELS:
            continue  # phone и пустые пропускаем
        st[k] = {"key": k, "name": r["name"], "channel": r["channel"],
                 "email": r.get("email", ""), "phone": r.get("phone", ""),
                 "telegram": r.get("telegram", ""), "vk": r.get("vk", ""),
                 "status": "new", "last_sent": "", "replied": ""}
    return st


def save_state(niche: str, st: dict):
    with state_path(niche).open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=STATE_COLS)
        w.writeheader()
        for r in st.values():
            w.writerow({c: r.get(c, "") for c in STATE_COLS})


def get_meta(niche):
    p = meta_path(niche)
    return json.loads(p.read_text()) if p.exists() else {}


def set_meta(niche, m):
    meta_path(niche).write_text(json.dumps(m, ensure_ascii=False))


def cap_for(cfg, day_num, ch):
    cap = cfg["caps"][ch]
    wu = cfg.get("warmup", {})
    if wu.get("enabled"):
        cap = min(cap, wu["start"] + wu["step"] * (day_num - 1))
    return max(cap, 0)


def days_since(d: str) -> int:
    dt = dparse(d)
    return (today() - dt).days if dt else 9999


def cmd_today(a):
    cfg = yaml.safe_load((CONFIG / "send_config.yaml").read_text(encoding="utf-8"))
    outreach = load_outreach(a.niche)
    st = load_state(a.niche, outreach)
    meta = get_meta(a.niche)
    start = dparse(meta.get("start_date")) or today()
    meta["start_date"] = start.isoformat()
    day_num = (today() - start).days + 1

    fu = cfg["follow_up"]
    batch = []
    for ch in ["telegram", "vk"]:   # ручные каналы; email — командой send-email
        cap = cap_for(cfg, day_num, ch)
        picks = []
        pool = [s for s in st.values() if s["channel"] == ch and not s["replied"]]
        # 1) фоллоу-апы по сроку
        for s in pool:
            if len(picks) >= cap:
                break
            if s["status"] == "contacted" and days_since(s["last_sent"]) >= fu["first_after_days"]:
                picks.append((s, "followup1"))
            elif s["status"] == "followup1" and days_since(s["last_sent"]) >= fu["second_after_days"]:
                picks.append((s, "followup2"))
        # 2) новые
        for s in pool:
            if len(picks) >= cap:
                break
            if s["status"] == "new":
                picks.append((s, "first"))
        # применяем
        for s, action in picks:
            o = outreach.get(s["key"], {})
            text = {"first": o.get("message", ""), "followup1": o.get("followup1", ""),
                    "followup2": o.get("followup2", "")}[action]
            batch.append({
                "action": action, "name": s["name"], "channel": ch,
                "contact": s.get(ch, ""), "subject": o.get("subject", "") if ch == "email" else "",
                "text": text, "done": "",
            })
            s["last_sent"] = today().isoformat()
            s["status"] = {"first": "contacted", "followup1": "followup1", "followup2": "closed"}[action]

    out = DATA / f"today_{a.niche}_{today().isoformat()}.csv"
    with out.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=["action", "name", "channel", "contact", "subject", "text", "done"])
        w.writeheader()
        for b in batch:
            w.writerow(b)
    save_state(a.niche, st)
    set_meta(a.niche, meta)

    import collections
    c = collections.Counter((b["channel"], b["action"]) for b in batch)
    print(f"день кампании №{day_num}. пачка на сегодня: {len(batch)} сообщений")
    for (ch, act), n in sorted(c.items()):
        print(f"  {ch:9} {act:10} {n}")
    print(f"файл: {out}")
    print("отправь эти сообщения руками, кто ответит — отметь: send_manager.py replied --name «...»")


def _find(st, name):
    return [s for s in st.values() if name.lower() in s["name"].lower()]


def cmd_mark(a, replied: bool):
    outreach = load_outreach(a.niche)
    st = load_state(a.niche, outreach)
    hit = _find(st, a.name)
    if not hit:
        print("не нашёл лид", file=sys.stderr); return
    for s in hit:
        if replied:
            s["replied"] = "1"; s["status"] = "replied"
        else:
            s["status"] = "skip"; s["replied"] = "1"
        print(f"  {'ответил' if replied else 'исключён'}: {s['name']}")
    save_state(a.niche, st)


def cmd_stats(a):
    outreach = load_outreach(a.niche)
    st = load_state(a.niche, outreach)
    import collections
    c = collections.Counter(s["status"] for s in st.values())
    total = len(st)
    repl = sum(1 for s in st.values() if s["status"] == "replied")
    contacted = sum(1 for s in st.values() if s["status"] in ("contacted", "followup1", "closed", "replied"))
    print(f"всего лидов в работе: {total}")
    for k, v in c.most_common():
        print(f"  {k:12} {v}")
    if contacted:
        print(f"отклик: {repl}/{contacted} = {repl*100//contacted}%")


def _due_for_channel(st, cfg, day_num, ch):
    fu = cfg["follow_up"]
    cap = cap_for(cfg, day_num, ch)
    pool = [s for s in st.values() if s["channel"] == ch and not s["replied"]]
    picks = []
    for s in pool:
        if len(picks) >= cap:
            break
        if s["status"] == "contacted" and days_since(s["last_sent"]) >= fu["first_after_days"]:
            picks.append((s, "followup1"))
        elif s["status"] == "followup1" and days_since(s["last_sent"]) >= fu["second_after_days"]:
            picks.append((s, "followup2"))
    for s in pool:
        if len(picks) >= cap:
            break
        if s["status"] == "new":
            picks.append((s, "first"))
    return picks


def cmd_send_email(a):
    cfg = yaml.safe_load((CONFIG / "send_config.yaml").read_text(encoding="utf-8"))
    ecfg_path = CONFIG / "email_config.yaml"
    if not ecfg_path.exists():
        print("нет configs/email_config.yaml — заполни данные почты Педро", file=sys.stderr); return
    ec = yaml.safe_load(ecfg_path.read_text(encoding="utf-8"))
    outreach = load_outreach(a.niche)
    st = load_state(a.niche, outreach)
    meta = get_meta(a.niche)
    start = dparse(meta.get("start_date")) or today()
    meta["start_date"] = start.isoformat()
    day_num = (today() - start).days + 1

    picks = _due_for_channel(st, cfg, day_num, "email")
    if not picks:
        print("на сегодня по email никого (лимит/прогрев/все обработаны)"); return

    print(f"день №{day_num}. к отправке по email: {len(picks)}" + (" [ПРОБНЫЙ ПРОГОН]" if a.dry_run else ""))
    server = None
    if not a.dry_run:
        if ec.get("use_ssl", True):
            server = smtplib.SMTP_SSL(ec["smtp_host"], int(ec["smtp_port"]), context=ssl.create_default_context())
        else:
            server = smtplib.SMTP(ec["smtp_host"], int(ec["smtp_port"])); server.starttls(context=ssl.create_default_context())
        server.login(ec["smtp_user"], ec["smtp_password"])

    sent = failed = 0
    for s, action in picks:
        o = outreach.get(s["key"], {})
        base_subj = o.get("subject", "Предложение")
        subj = base_subj if action == "first" else "Re: " + base_subj
        body = {"first": o.get("message", ""), "followup1": o.get("followup1", ""),
                "followup2": o.get("followup2", "")}[action]
        body += "\n\n---\n" + ec.get("unsubscribe_text", "Если предложение не актуально — ответьте «отписаться», и я больше не напишу.")
        to = s.get("email", "")
        if not to:
            continue
        if a.dry_run:
            print(f"  [dry] {action:9} -> {to} | тема: {subj}")
            sent += 1
            continue
        try:
            msg = MIMEText(body, "plain", "utf-8")
            msg["Subject"] = subj
            msg["From"] = formataddr((ec.get("from_name", ""), ec["from_email"]))
            msg["To"] = to
            server.sendmail(ec["from_email"], [to], msg.as_string())
            s["last_sent"] = today().isoformat()
            s["status"] = {"first": "contacted", "followup1": "followup1", "followup2": "closed"}[action]
            sent += 1
            print(f"  отправлено {action:9} -> {to}")
            time.sleep(float(ec.get("send_delay_sec", 20)))
        except Exception as exc:
            failed += 1
            print(f"  ! ошибка {to}: {exc}", file=sys.stderr)
    if server:
        server.quit()
    if not a.dry_run:
        save_state(a.niche, st); set_meta(a.niche, meta)
    print(f"итог: отправлено {sent}, ошибок {failed}")


def cmd_dashboard(a):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("нужен openpyxl: python -m pip install openpyxl", file=sys.stderr); return
    outreach = load_outreach(a.niche)
    st = load_state(a.niche, outreach)
    wb = Workbook()

    ws = wb.active; ws.title = "Лиды"
    headers = ["Название", "Канал", "Контакт", "Статус", "Последнее касание", "Ответил"]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="2B2B2B")
    colors = {"new": "FFFFFF", "contacted": "FFF2CC", "followup1": "FCE5CD",
              "followup2": "FCE5CD", "closed": "EAD1DC", "replied": "D9EAD3", "skip": "EEEEEE"}
    for s in sorted(st.values(), key=lambda x: x["status"]):
        row = [s["name"], s["channel"], s.get(s["channel"], ""), s["status"], s.get("last_sent", ""),
               "да" if s.get("replied") else ""]
        ws.append(row)
        fill = colors.get(s["status"], "FFFFFF")
        for c in ws[ws.max_row]:
            c.fill = PatternFill("solid", fgColor=fill)
    for col, w in zip("ABCDEF", (32, 12, 34, 14, 18, 10)):
        ws.column_dimensions[col].width = w

    wf = wb.create_sheet("Воронка")
    import collections
    cnt = collections.Counter(s["status"] for s in st.values())
    contacted = sum(cnt[k] for k in ("contacted", "followup1", "closed", "replied"))
    repl = cnt.get("replied", 0)
    wf.append(["Метрика", "Значение"])
    wf["A1"].font = wf["B1"].font = Font(bold=True)
    wf.append(["Всего лидов", len(st)])
    for k in ("new", "contacted", "followup1", "followup2", "closed", "replied", "skip"):
        if cnt.get(k):
            wf.append([k, cnt[k]])
    wf.append(["Отклик, %", f"{(repl*100//contacted) if contacted else 0}%"])
    wf.column_dimensions["A"].width = 20; wf.column_dimensions["B"].width = 12

    out = DATA / f"dashboard_{a.niche}.xlsx"
    wb.save(out)
    print(f"дашборд готов: {out}")


def main():
    ap = argparse.ArgumentParser()
    sub = ap.add_subparsers(dest="cmd")
    for name in ("today", "stats", "dashboard"):
        p = sub.add_parser(name); p.add_argument("--niche", required=True)
    pe = sub.add_parser("send-email"); pe.add_argument("--niche", required=True)
    pe.add_argument("--dry-run", action="store_true", help="показать кому уйдёт, но не отправлять")
    for name in ("replied", "skip"):
        p = sub.add_parser(name); p.add_argument("--niche", required=True); p.add_argument("--name", required=True)
    a = ap.parse_args()
    if a.cmd == "today": cmd_today(a)
    elif a.cmd == "stats": cmd_stats(a)
    elif a.cmd == "dashboard": cmd_dashboard(a)
    elif a.cmd == "send-email": cmd_send_email(a)
    elif a.cmd == "replied": cmd_mark(a, True)
    elif a.cmd == "skip": cmd_mark(a, False)
    else: ap.print_help()


if __name__ == "__main__":
    main()
